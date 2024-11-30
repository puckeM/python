import telebot
import openpyxl
import os
import logging
from PIL import Image
import schedule
import time
import threading
from datetime import datetime


#файловая система
from bot.materials import def_get_text

# настройка ведения логов
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',level=logging.INFO)


bot = telebot.TeleBot('7324666918:AAENCFrgssoHCMVBCBnOmT1SH1Nz1H49YR0')

# расписание
training_schedule = {
    'Monday': ['18:00'],
    'Wednesday': ['18:00'],
    'Friday': ['18:00']
}


# создание файла excel и также проверка существует ли он или нет
if not os.path.exists('player_data.xlsx'):
    workbook = openpyxl.Workbook()

    sheet = workbook.active
    sheet['A1'] = 'Имя'
    sheet['B1'] = 'Вес'
    sheet['C1'] = 'Рост'
    sheet['D1'] = 'Факультет'
    sheet['E1'] = 'Любимые книги'
    sheet['F1'] = 'Любимый регбийный клуб'

    for column in range(1, 7):
        sheet.column_dimensions[chr(column + 64)].width = 30

else:
    workbook = openpyxl.load_workbook('player_data.xlsx')
    sheet = workbook.active


# глобальные переменные
player_data = {}
admin_mode = False
player_registered = {}
user_hi = False
open_photo = True

# флаги для обработчиков продолжения
flag_schedule = False

# функции

# функции связаны с тренировками
def training_reminder(chat_id): #оповещение о тренировках
    current_day = datetime.now().strftime('%A')
    current_time = datetime.now().strftime('%H:%M')

    if current_time in training_schedule.get(current_day, []):
        bot.send_message(chat_id,f"Внимание! Сегодня тренировка в {current_time}! Пожалуйста, подтвердите свое присутствие.")

    markup = telebot.types.ReplyKeyboardMarkup(one_time_keyboard=True)
    b1 = telebot.types.KeyboardButton('✅')
    b2 = telebot.types.KeyboardButton('❌')
    markup.add(b1, b2)

    bot.send_message(chat_id, 'Вы сможете прийти на тренировку?', reply_markup=markup)

def end_training(chat_id):
    time.sleep(3600)  #тренировка длится 1 час, значение в секундах
    summary = "Тренировка завершена! Вот краткое содержание:\n- Разминка\n- Основная часть\n- Заминка"
    bot.send_message(chat_id, summary)

def schedule_trainings():
    for chat_id in player_registered.keys():
        training_reminder(chat_id)
        threading.Thread(target=end_training, args=(chat_id,)).start()

def job():
    current_day = datetime.now().strftime('%A')
    current_time = datetime.now().strftime('%H:%M')

    if current_time in training_schedule.get(current_day, []):
        schedule_trainings()

schedule.every(1).minutes.do(job)

def process_schedule(message):
    global training_schedule

    try:
        # Разделяем введенное сообщение по ';'
        days = message.text.split(';')
        new_schedule = {}

        for day in days:
            parts = day.split()
            day_name = parts[0].strip()
            times = [time.strip() for time in parts[1].split(',')]

            # Обновляем расписание
            new_schedule[day_name] = times

        training_schedule = new_schedule
        bot.send_message(message.chat.id, f"Новое расписание установлено:\n{training_schedule}")
        verify_admin(message, 'admin')


    except Exception as e:
        bot.send_message(message.chat.id, "Ошибка! Убедитесь, что вы ввели данные в правильном формате.")
        print(e)
        new_schedule(message)
#

'''def set_schedule(message):
    global admin_mode, fail_reg, flag_exit

    if admin_mode and not fail_reg:
        markup = telebot.types.InlineKeyboardMarkup()

        b1 = telebot.types.InlineKeyboardButton('exit', callback_data='exit')
        markup.add(b1)

        bot.send_message(message.chat.id, "Введите новое расписание в формате 'день1 время1, время2; день2 время1, время2'. Пример: 'Monday 18:00, 19:00; Wednesday 18:00'.",
                         reply_markup=markup)
        # Устанавливаем состояние ожидания нового расписания

        print(training_schedule)

        if flag_exit:
            verify_admin(message, 'admin')
            flag_exit = False

        else:
            get_process_schedule(message)

    elif admin_mode and fail_reg:
        bot.send_message(message.chat.id, "Введите ёще раз новое расписание в формате 'день1 время1, время2; день2 время1, время2'. Пример: 'Monday 18:00, 19:00; Wednesday 18:00'.")
        print(training_schedule)
        fail_reg = False
        get_process_schedule(message)'''
# разбиение функции set_schedule
def ask_for_continue(message):
    global flag_schedule

    flag_schedule = True
    bot.send_message(message.chat.id, 'Вы хотите продолжить?', reply_markup=return_to_back())

def new_schedule(message):
    global admin_mode

    if admin_mode:
        bot.send_message(message.chat.id, "Введите новое расписание в формате 'день1 время1, время2; день2 время1, время2'. Пример: 'Monday 18:00, 19:00; Wednesday 18:00'.")
        bot.register_next_step_handler(message, process_schedule)
    else:
        bot.send_message(message.chat.id, 'Неверная авторизация')
        start(message)


#

'''def exit():
    global flag_exit

    markup = telebot.types.InlineKeyboardMarkup()
    flag_exit = True

    b1 = telebot.types.InlineKeyboardButton('exit', callback_data='exit')
    markup.add(b1)

    return markup'''

def keyboard_start():
    markup = telebot.types.ReplyKeyboardMarkup(row_width=2)

    b1 = telebot.types.KeyboardButton('Админ')
    b2 = telebot.types.KeyboardButton('📄')
    b3 = telebot.types.KeyboardButton('Расписание1')
    b4 = telebot.types.KeyboardButton('Матер')

    markup.add(b1, b2, b3, b4)
    return markup

def return_to_back():
    markup = telebot.types.InlineKeyboardMarkup()
    b1 = telebot.types.InlineKeyboardButton('✅', callback_data='yes_continue')
    b2 = telebot.types.InlineKeyboardButton('❌', callback_data='no_continue')
    markup.add(b1, b2)

    return markup

def create_keyboard():
    markup = telebot.types.InlineKeyboardMarkup()
    item1 = telebot.types.InlineKeyboardButton('▶', callback_data='start_func')
    item2 = telebot.types.InlineKeyboardButton('❔', callback_data='help')
    markup.add(item1, item2)
    return markup

def create_keybord_admin():
    markup = telebot.types.InlineKeyboardMarkup()
    item1 = telebot.types.InlineKeyboardButton('▶', callback_data='start_func')
    item2 = telebot.types.InlineKeyboardButton('❔', callback_data='help')
    markup.add(item1, item2)
    return markup

def create_keybord_on_message():
    b1 = telebot.types.InlineKeyboardButton('1', callback_data='material1')
    b2 = telebot.types.InlineKeyboardButton('2', callback_data='material2')
    b3 = telebot.types.InlineKeyboardButton('3', callback_data='material3')
    b4 = telebot.types.InlineKeyboardButton('4', callback_data='material4')
    b5 = telebot.types.InlineKeyboardButton('5', callback_data='material5')
    b6 = telebot.types.InlineKeyboardButton('6', callback_data='material6')
    b7 = telebot.types.InlineKeyboardButton('7', callback_data='material7')
    b8 = telebot.types.InlineKeyboardButton('8', callback_data='material8')
    b9 = telebot.types.InlineKeyboardButton('9', callback_data='material9')
    b10 = telebot.types.InlineKeyboardButton('10', callback_data='materialA')

    list1_5 = [b1, b2, b3, b4, b5]
    list6_10 = [b6, b7, b8, b9, b10]

    buttons = [list1_5, list6_10]

    markup = telebot.types.InlineKeyboardMarkup(buttons)

    return markup

# методички
def material(x):
    #file = f'materials/text{x}/1_1.jpg'
    txt = ''
    i = 0
    with open(f'materials/text{x}/text.txt', 'r', encoding='utf-8') as obj:
        if x == '2':
            for line in obj.readlines():
                i += 1
                if i == 1:
                    txt += line + '\n'
                else:
                    txt += line
            i = 0
        else:
            for line in obj.readlines():
                txt += line

        return txt


def photos(x):
    global open_photo

    if open_photo:
        file1 = Image.open(f'materials/text{x}/1_1.jpg')
        open_photo = False
        return file1

    else:
        file2 = Image.open(f'materials/text{x}/1_2.jpg')
        open_photo = True

        return file2

# обработчик опоповещий о тренировках

@bot.message_handler(func=lambda message: message.text == 'Расписание1')
def handle_message(message):
    chat_id = message.chat.id

    if message.text == '✅':
        player_registered[chat_id] = 'confirmed'
        bot.send_message(chat_id, "Спасибо за подтверждение! Ждем вас на тренировке.")
    elif message.text == '❌':
        player_registered[chat_id] = 'not_confirmed'
        bot.send_message(chat_id, 'Поняли, что вы не сможете прийти. Увидимся в следующий раз.')
    else:
        bot.send_message(chat_id, "Пожалуйста, подтвердите свое присутствие на тренировке.")

#
@bot.message_handler(commands=['start'])
def start(message): # начальная функция, при вызове которой вызываются кнопки-функционала
    global admin_mode, user_hi
    admin_mode = False # сделано для того чтобы работала функция выхода из админ режима

    # процесс создания кнопок

    #item3 = telebot.types.KeyboardButton('Админ')
    #item4 = telebot.types.KeyboardButton('📄')
    #item5 = telebot.types.KeyboardButton('Матер')
    #item6 = telebot.types.KeyboardButton('Расписание1')

    if not user_hi:
        bot.send_message(message.chat.id, 'Привет', reply_markup=keyboard_start())

    bot.send_message(message.chat.id, 'Выберите действие:', reply_markup=create_keyboard()) # reply_marlup нужен для вызова тех самых кнопок

@bot.message_handler(func=lambda message: message.text == 'Расписание1')
def get_schedule_users(message):
    global training_schedule
    bot.send_message(message.chat.id, f'Текущее расписание тренировок:\n'
                                      f'{training_schedule}')
    start(message)

@bot.message_handler(func=lambda message: message.text == 'Матер')
def get_materials(message):
    bot.send_message(message.chat.id, '1 - захват asdlkasldkasdka;m asdlkasldk;asd\n'
                                      '2 - asdlasldkasdk ;lkad;la kslkd asdkasdkas;dlkl\n'
                                      '3 - asdlasldkasdk ;lkad;la kslkd asdkasdkas;dlkl\n'
                                      '4 - asdlasldkasdk ;lkad;la kslkd asdkasdkas;dlkl\n'
                                      '5 - asdlasldkasdk ;lkad;la kslkd asdkasdkas;dlkl\n'
                                      '6 - asdlasldkasdk ;lkad;la kslkd asdkasdkas;dlkl\n'
                                      '7 - asdlasldkasdk ;lkad;la kslkd asdkasdkas;dlkl\n'
                                      '8 - asdlasldkasdk ;lkad;la kslkd asdkasdkas;dlkl\n'
                                      '9 - asdlasldkasdk ;lkad;la kslkd asdkasdkas;dlkl\n'
                                      '10 - asdlasldkasdk ;lkad;la kslkd asdkasdkas;dlkl\n', reply_markup=create_keybord_on_message())

#выдача материала
@bot.callback_query_handler(func=lambda call: call.data == 'material1')
def get_material(call):
    global user_hi

    user_hi = True
    message = call.message
    x = call.data[-1]

    bot.send_media_group(message.chat.id, media=[telebot.types.InputMediaPhoto(photos(x)), telebot.types.InputMediaPhoto(photos(x), caption=material(x))])

    bot.send_message(message.chat.id, 'Желаетe ещё что-то узнать?', reply_markup=return_to_back())

@bot.callback_query_handler(func=lambda call: call.data == 'material2')
def get_material(call):
    global user_hi

    user_hi = True
    message = call.message
    x = call.data[-1]

    bot.send_media_group(message.chat.id, media=[telebot.types.InputMediaPhoto(photos(x)), telebot.types.InputMediaPhoto(photos(x), caption=material(x))])

    bot.send_message(message.chat.id, 'Желаетe ещё что-то узнать?', reply_markup=return_to_back())

@bot.callback_query_handler(func=lambda call: call.data == 'material3')
def get_material(call):
    global user_hi

    user_hi = True
    message = call.message
    x = call.data[-1]

    bot.send_media_group(message.chat.id, media=[telebot.types.InputMediaPhoto(photos(x)), telebot.types.InputMediaPhoto(photos(x), caption=material(x))])

    bot.send_message(message.chat.id, 'Желаетe ещё что-то узнать?', reply_markup=return_to_back())

@bot.callback_query_handler(func=lambda call: call.data == 'material4')
def get_material(call):
    global user_hi

    user_hi = True
    message = call.message
    x = call.data[-1]

    bot.send_media_group(message.chat.id, media=[telebot.types.InputMediaPhoto(photos(x)), telebot.types.InputMediaPhoto(photos(x), caption=material(x))])

    bot.send_message(message.chat.id, 'Желаетe ещё что-то узнать?', reply_markup=return_to_back())

@bot.callback_query_handler(func=lambda call: call.data == 'material5')
def get_material(call):
    global user_hi

    user_hi = True
    message = call.message
    x = call.data[-1]

    bot.send_media_group(message.chat.id, media=[telebot.types.InputMediaPhoto(photos(x)), telebot.types.InputMediaPhoto(photos(x), caption=material(x))])

    bot.send_message(message.chat.id, 'Желаетe ещё что-то узнать?', reply_markup=return_to_back())

@bot.callback_query_handler(func=lambda call: call.data == 'material6')
def get_material(call):
    global user_hi

    user_hi = True
    message = call.message
    x = call.data[-1]

    bot.send_media_group(message.chat.id, media=[telebot.types.InputMediaPhoto(photos(x)), telebot.types.InputMediaPhoto(photos(x), caption=material(x))])

    bot.send_message(message.chat.id, 'Желаетe ещё что-то узнать?', reply_markup=return_to_back())

@bot.callback_query_handler(func=lambda call: call.data == 'material7')
def get_material(call):
    global user_hi

    user_hi = True
    message = call.message
    x = call.data[-1]

    bot.send_media_group(message.chat.id, media=[telebot.types.InputMediaPhoto(photos(x)), telebot.types.InputMediaPhoto(photos(x), caption=material(x))])

    bot.send_message(message.chat.id, 'Желаетe ещё что-то узнать?', reply_markup=return_to_back())

@bot.callback_query_handler(func=lambda call: call.data == 'material8')
def get_material(call):
    global user_hi

    user_hi = True
    message = call.message
    x = call.data[-1]

    bot.send_media_group(message.chat.id, media=[telebot.types.InputMediaPhoto(photos(x)), telebot.types.InputMediaPhoto(photos(x), caption=material(x))])

    bot.send_message(message.chat.id, 'Желаетe ещё что-то узнать?', reply_markup=return_to_back())

@bot.callback_query_handler(func=lambda call: call.data == 'material9')
def get_material(call):
    global user_hi

    user_hi = True
    message = call.message
    x = call.data[-1]

    bot.send_media_group(message.chat.id, media=[telebot.types.InputMediaPhoto(photos(x)), telebot.types.InputMediaPhoto(photos(x), caption=material(x))])

    bot.send_message(message.chat.id, 'Желаетe ещё что-то узнать?', reply_markup=return_to_back())

@bot.callback_query_handler(func=lambda call: call.data == 'materialA')
def get_material(call):
    global user_hi

    user_hi = True
    message = call.message
    x = call.data[-1]

    bot.send_media_group(message.chat.id, media=[telebot.types.InputMediaPhoto(photos(x)), telebot.types.InputMediaPhoto(photos(x), caption=material(x))])

    bot.send_message(message.chat.id, 'Желаетe ещё что-то узнать?', reply_markup=return_to_back())

# end

@bot.callback_query_handler(func=lambda call: call.data == 'yes_continue')
def return_to(call):
    global flag_schedule
    message = call.message

    if flag_schedule:
        new_schedule(message)

    else:
        get_materials(message)

@bot.callback_query_handler(func=lambda call: call.data == 'no_continue')
def return_back(call):
    global flag_schedule
    message = call.message

    if flag_schedule:
        verify_admin(message)
    else:
        start(message)

@bot.callback_query_handler(func=lambda call: call.data == 'help')
def help(call):
    global admin_mode, user_hi

    #user_hi = True
    message = call.message
    if not admin_mode:
        bot.send_message(message.chat.id, '▶ - прохождение регистрации пользователя\n'
                                          '❔ - помощь в использовании бота\n'
                                          '📄 - изменить данные о себе')

        start(message)
    elif admin_mode:
        bot.send_message(message.chat.id, '▶ - прохождение регистрации пользователя\n'
                                          '❔ - помощь в использовании бота\n'
                                          '📄 - изменить данные о себе\n'
                                          'ADMIN CHECK')

        verify_admin(message, 'admin')


# при выводе сообщения старт, вызывается функция collect_data
@bot.callback_query_handler(func=lambda call: call.data == 'start_func')
def collect_data(call):
    global admin_mode, user_hi

    user_hi = True
    message = call.message
    user_ip = message.chat.id # для проверки регистрации

    if user_ip not in player_registered:
        bot.send_message(message.chat.id, 'Введите ваше имя:')
        bot.register_next_step_handler(message, get_weight)
    elif admin_mode:
        bot.send_message(message.chat.id, 'Введите имя игрока:')
        bot.register_next_step_handler(message, get_weight)
    else:
        bot.send_message(message.chat.id, 'Вы уже зарегистрированы. Что вы хотите изменить?')
        bot.register_next_step_handler(message, edit_data)

# процесс сбора информации об игроке, каждая функция отвечает за каждую характеристику игрока
def get_weight(message):
    global admin_mode

    # проверка: если пользователь вошёл в режим админа, значит у него есть возможность добавлять любого игрока в базу
    # если пользователь не в режиме админа, значит он проходит первую регистрацию

    if admin_mode:
        # добавление информации в базу
        row = sheet.max_row
        sheet.cell(row=row, column=1, value=message.text.title())

    # если пользователь не в режиме админа, значит он проходит первую регистрацию

    else:
        player_data['name'] = message.text.title()

    bot.send_message(message.chat.id, 'Введите ваш вес')
    bot.register_next_step_handler(message, get_height) # эта строка отвечает за то, чтобы после запуска функции
                                                        # сразу же за ней следовала функция которая описана в этом методе

# дальше аналогично и первой функции
def get_height(message):
    global admin_mode

    if admin_mode:
        row = sheet.max_row
        sheet.cell(row=row, column=2, value=message.text)
    else:
        player_data['weight'] = message.text

    bot.send_message(message.chat.id, "Введите ваш рост:")
    bot.register_next_step_handler(message, get_faculty)

def get_faculty(message):
    global admin_mode

    if admin_mode:
        row = sheet.max_row
        sheet.cell(row=row, column=3, value=message.text)
    else:
        player_data['height'] = message.text

    bot.send_message(message.chat.id, 'Введите ваш факультет:')
    bot.register_next_step_handler(message, get_books)

def get_books(message):
    global admin_mode

    if admin_mode:
        row = sheet.max_row
        sheet.cell(row=row, column=4, value=message.text)
    else:
        player_data['faculty'] = message.text

    bot.send_message(message.chat.id, 'Введите ваши любимые книги:')
    bot.register_next_step_handler(message, get_club)

def get_club(message):
    global admin_mode

    if admin_mode:
        row = sheet.max_row
        sheet.cell(row=row, column=5, value=message.text)
    else:
        player_data['books'] = message.text.title()

    bot.send_message(message.chat.id, 'Введите ваш любимый регбийный клуб:')
    bot.register_next_step_handler(message, save_data)

def save_data(message):
    global admin_mode

    user_ip = message.chat.id

    if admin_mode:
        row = sheet.max_row
        sheet.cell(row=row, column=6, value=message.text)
    else:
        player_data['club'] = message.text.title()
    # сохранениие всеё информации в таблицу
    row = [player_data['name'], player_data['weight'], player_data['height'], player_data['faculty'], player_data['books'], player_data['club']]
    sheet.append(row)
    workbook.save('player_data.xlsx')

    if admin_mode:
        player_registered[user_ip] = False
    else:
        player_registered[user_ip] = True
    bot.send_message(message.chat.id, 'Данные сохранены!')
    bot.send_message(message.chat.id, text='Выберите действие:', reply_markup=create_keyboard())

@bot.message_handler(func=lambda message: message.text == 'Файл')
def send_file(message):
    if admin_mode:
        workbook.save('player_data.xlsx')
        with open('player_data.xlsx', 'rb') as file:
            bot.send_document(message.chat.id, file)
    else:
        bot.send_message(message.chat.id, 'Доступ запрещен. Войдите в режим админа.')

#@bot.callback_query_handler(func=lambda call: call.data == 'start_admin')
@bot.message_handler(func=lambda message: message.text == 'Админ')
def admin_mode(message):
    global admin_mode

    bot.send_message(message.chat.id, 'Введите логин:')
    bot.register_next_step_handler(message, admin_password)

def admin_password(message):
    login = message.text
    bot.send_message(message.chat.id, 'Введите пароль:')
    bot.register_next_step_handler(message, verify_admin, login)

def verify_admin(message, login):
    global admin_mode, user_hi, admin_hi

    user_hi = True
    password = message.text
    if (login == 'admin' and password == 'password') or admin_mode:
        if admin_mode == True:
            markup = telebot.types.ReplyKeyboardMarkup(row_width=2)
            #item1 = telebot.types.KeyboardButton('Старт')
            #item2 = telebot.types.KeyboardButton('Файл')
            item3 = telebot.types.KeyboardButton('Выход')
            item4 = telebot.types.KeyboardButton('Изм_игрока')
            item5 = telebot.types.KeyboardButton('Расписание')

            markup.add(item3, item4, item5)

            bot.send_message(message.chat.id, 'Вы в меню админа', reply_markup=markup)
            bot.send_message(message.chat.id, 'Выбери действие:', reply_markup=create_keyboard())

            bot.register_next_step_handler(message, admin_actions)
        else:
            admin_mode = True
            markup = telebot.types.ReplyKeyboardMarkup(row_width=2)
            #item1 = telebot.types.KeyboardButton('Старт')
            #item2 = telebot.types.KeyboardButton('Файл')
            item3 = telebot.types.KeyboardButton('Выход')
            item4 = telebot.types.KeyboardButton('Изм_игрока')
            item5 = telebot.types.KeyboardButton('Расписание')
            markup.add(item3, item4, item5)

            bot.send_message(message.chat.id, 'Добро пожаловать в режим админа!', reply_markup=markup)
            bot.send_message(message.chat.id, 'Выбери действие:', reply_markup=create_keyboard())

            bot.register_next_step_handler(message, admin_actions)
    else:
        bot.send_message(message.chat.id, 'Неверный логин или пароль.')

@bot.callback_query_handler(func=lambda call: call.data == 'exit')
def exit(call):
    global flag_exit
    message = call.message
    flag_exit = True
    verify_admin(message, 'admin')


def admin_actions(message):
    global admin_mode
    if message.text == 'Старт':
        collect_data(message)
    elif message.text == 'Файл':
        send_file(message)
    elif message.text == 'Выход':
        admin_mode = False
        bot.send_message(message.chat.id, 'Ты вышел из режима админа', reply_markup=keyboard_start())
        start(message)
    elif message.text == 'Изм_игрока':
        admin_edit(message)
    elif message.text == 'Расписание':
        ask_for_continue(message)

@bot.message_handler(func=lambda message: message.text == 'Изм_игрока')
def admin_edit(message):
    global admin_mode
    if admin_mode:
        workbook = openpyxl.load_workbook('player_data.xlsx')
        worksheet = workbook.active
        player_names = []
        for row in range(2, worksheet.max_row + 1):
            player_names.append(worksheet.cell(row=row, column=1).value)
        keyboard = telebot.types.ReplyKeyboardMarkup(row_width=2)
        for name in player_names:
            keyboard.add(telebot.types.KeyboardButton(name))
        keyboard.add(telebot.types.KeyboardButton('назад'))

        bot.send_message(chat_id=message.chat.id, text='Выберите игрока, которого хотите изменить:',
                         reply_markup=keyboard)
        bot.register_next_step_handler(message, admin_edit_player_data)
    else:
        bot.send_message(chat_id=message.chat.id, text='Вы не являетесь админом.')

def choose(message):
    if message.text == 'назад':
        verify_admin(message, 'admin')


def admin_edit_player_data(message):
    global player_to_edit
    player_to_edit = message.text

    workbook = openpyxl.load_workbook('player_data.xlsx')
    worksheet = workbook.active

    if message.text == 'назад':
        verify_admin(message, 'admin')
    else:
        for row in range(2, worksheet.max_row + 1):
            if worksheet.cell(row=row, column=1).value == player_to_edit:
                bot.send_message(chat_id=message.chat.id, text=f'Текущие данные игрока {player_to_edit}:\n'
                                                               f'Вес: {worksheet.cell(row=row, column=2).value}\n'
                                                               f'Рост: {worksheet.cell(row=row, column=3).value}\n'
                                                               f'Факультет: {worksheet.cell(row=row, column=4).value}\n'
                                                               f'Любимые книги: {worksheet.cell(row=row, column=5).value}\n'
                                                               f'Любимый регбийный клуб: {worksheet.cell(row=row, column=6).value}')

                bot.send_message(chat_id=message.chat.id, text='Что вы хотите изменить?')
                keyboard = telebot.types.ReplyKeyboardMarkup(row_width=2)
                keyboard.add(telebot.types.KeyboardButton('Имя'),
                             telebot.types.KeyboardButton('Вес'),
                             telebot.types.KeyboardButton('Рост'),
                             telebot.types.KeyboardButton('Факультет'),
                             telebot.types.KeyboardButton('Любимые книги'),
                             telebot.types.KeyboardButton('Любимый регбийный клуб'),
                             telebot.types.KeyboardButton('назад'))
                bot.send_message(chat_id=message.chat.id, text='Выберите поле для изменения:', reply_markup=keyboard)
                bot.register_next_step_handler(message, admin_update_player_data)
                break
        else:
            bot.send_message(chat_id=message.chat.id, text='Игрок не найден.')


def admin_update_player_data(message):
    field_to_update = message.text

    bot.send_message(chat_id=message.chat.id, text=f'Введите новое значение для поля "{field_to_update}"')
    bot.register_next_step_handler(message, lambda m: admin_save_player_data(m, field_to_update))

def admin_save_player_data(message, field_to_update):
    new_value = message.text

    workbook = openpyxl.load_workbook('player_data.xlsx')
    worksheet = workbook.active

    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == player_to_edit:
            if field_to_update == 'Имя':
                worksheet.cell(row=row, column=1, value=new_value.title())
                verify_admin(message, 'admin')
            elif field_to_update == 'Вес':
                worksheet.cell(row=row, column=2, value=new_value)
                verify_admin(message, 'admin')
            elif field_to_update == 'Рост':
                worksheet.cell(row=row, column=3, value=new_value)
                verify_admin(message, 'admin')
            elif field_to_update == 'Факультет':
                worksheet.cell(row=row, column=4, value=new_value)
                verify_admin(message, 'admin')
            elif field_to_update == 'Любимые книги':
                worksheet.cell(row=row, column=5, value=new_value.title())
                verify_admin(message, 'admin')
            elif field_to_update == 'Любимый регбийный клуб':
                worksheet.cell(row=row, column=6, value=new_value.title())
                verify_admin(message, 'admin')
            elif field_to_update == 'назад':
                verify_admin(message, 'admin')
            workbook.save('player_data.xlsx')
            bot.send_message(chat_id=message.chat.id, text=f'Данные игрока {player_to_edit} успешно обновлены.')
            break
    else:
        bot.send_message(chat_id=message.chat.id, text='Игрок не найден.')


@bot.message_handler(func=lambda message: message.text == '📄')
def edit_data(message):
    user_ip = message.chat.id
    if user_ip in player_registered:
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == player_data['name']:
                bot.send_message(message.chat.id, f'ФИО: {sheet.cell(row=row, column=1).value}')
                bot.send_message(message.chat.id, f"Вес: {sheet.cell(row=row, column=2).value}")
                bot.send_message(message.chat.id, f"Рост: {sheet.cell(row=row, column=3).value}")
                bot.send_message(message.chat.id, f"Факультет: {sheet.cell(row=row, column=4).value}")
                bot.send_message(message.chat.id, f"Любимые книги: {sheet.cell(row=row, column=5).value}")
                bot.send_message(message.chat.id, f"Любимый регбийный клуб: {sheet.cell(row=row, column=6).value}")
                markup = telebot.types.ReplyKeyboardMarkup(row_width=2)
                item1 = telebot.types.KeyboardButton('ФИО')
                item2 = telebot.types.KeyboardButton('Вес')
                item3 = telebot.types.KeyboardButton('Рост')
                item4 = telebot.types.KeyboardButton('Факультет')
                item5 = telebot.types.KeyboardButton('Любимые книги')
                item6 = telebot.types.KeyboardButton('Любимый регбийный клуб')
                item7 = telebot.types.KeyboardButton('Назад')
                markup.add(item1, item2, item3, item4, item5, item6, item7)
                bot.send_message(message.chat.id, "Что вы хотите изменить?", reply_markup=markup)
                bot.register_next_step_handler(message, update_data)
                break
    else:
        bot.send_message(message.chat.id, 'Вы еще не зарегистрированы. Пройдите регистрацию.')
        start(message)

def update_data(message):
    if message.text == 'ФИО':
        bot.send_message(message.chat.id, 'Введите новые ФИО:')
        bot.register_next_step_handler(message, update_name)
    elif message.text == 'Вес':
        bot.send_message(message.chat.id, 'Введите новый вес:')
        bot.register_next_step_handler(message, update_weight)
    elif message.text == 'Рост':
        bot.send_message(message.chat.id, 'Введите новый рост:')
        bot.register_next_step_handler(message, update_height)
    elif message.text == 'Факультет':
        bot.send_message(message.chat.id, 'Введите новый факультет:')
        bot.register_next_step_handler(message, update_faculty)
    elif message.text == 'Любимые книги':
        bot.send_message(message.chat.id, 'Введите новые любимые книги:')
        bot.register_next_step_handler(message, update_books)
    elif message.text == 'Любимый регбийный клуб':
        bot.send_message(message.chat.id, 'Введите новый любимый регбийный клуб:')
        bot.register_next_step_handler(message, update_club)
    elif message.text == 'Назад':
        start(message)
    else:
        bot.send_message(message.chat.id, 'Неверный выбор. Попробуйте еще раз.')
        edit_data(message)

def update_name(message):
    player_data['name'] = message.text.title()
    update_row(message)

def update_weight(message):
    player_data['weight'] = message.text
    update_row(message)

def update_height(message):
    player_data['height'] = message.text
    update_row(message)

def update_faculty(message):
    player_data['faculty'] = message.text
    update_row(message)

def update_books(message):
    player_data['books'] = message.text.title()
    update_row(message)

def update_club(message):
    player_data['club'] = message.text.title()
    update_row(message)

def update_row(message):
    row_num = None
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == player_data['name']:
            row_num = row
            break
    if row_num:
        sheet.cell(row=row_num, column=1, value=player_data['name'])
        sheet.cell(row=row_num, column=2, value=player_data['weight'])
        sheet.cell(row=row_num, column=3, value=player_data['height'])
        sheet.cell(row=row_num, column=4, value=player_data['faculty'])
        sheet.cell(row=row_num, column=5, value=player_data['books'])
        sheet.cell(row=row_num, column=6, value=player_data['club'])
        workbook.save('player_data.xlsx')
        bot.send_message(message.chat.id, 'Данные обновлены.')
        start(message)
    else:
        bot.send_message(message.chat.id, 'Игрок не найден.')
        start(message)

@bot.message_handler(func=lambda message: True)
def cancel(message):
    bot.send_message(message.chat.id, 'Отмена. Введите /start, чтобы начать заново.')


def main():
    while True:

        schedule.run_pending()
        time.sleep(1)

        try:
            threading.Thread(bot.polling(none_stop=True)).start()
            #bot.polling()
        except Exception as E:
            logging.error(f'ошибка в работе алгоритма: {E}')
            continue


if __name__ == '__main__':
    main()